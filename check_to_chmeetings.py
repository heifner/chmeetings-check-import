"""
Check Image PDF -> CHMeetings Contribution Import  (v3)
=======================================================
Extracts contribution data from scanned check images in a PDF
and outputs a CHMeetings-compatible Excel file.

All processing happens locally on your machine -- nothing is uploaded.

Requirements (install once):
    pip install pymupdf easyocr openpyxl Pillow

Usage:
    python check_to_chmeetings.py "checks.pdf" --review
    python check_to_chmeetings.py "checks.pdf" --review --fund "Tithes" --date "03/08/2026"
    python check_to_chmeetings.py "checks.pdf" --review --batch "March 2026" --deposit-date "03/08/2026"

In review mode, each check image is displayed in a window so you can read
the check while confirming or correcting the OCR suggestions.
"""

import argparse
import difflib
import json
import os
import re
import sys
import threading
from datetime import datetime
from pathlib import Path

try:
    import fitz  # PyMuPDF
except ImportError:
    sys.exit("Missing dependency: pip install pymupdf")

try:
    import easyocr
except ImportError:
    sys.exit("Missing dependency: pip install easyocr")

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

from PIL import Image, ImageEnhance, ImageFilter, ImageTk
import io
import numpy as np

try:
    import tkinter as tk
    HAS_TK = True
except ImportError:
    HAS_TK = False


# -- Configuration -----------------------------------------------------------

CHMEETINGS_COLUMNS = [
    "First Name", "Last Name", "Envelope Number", "Email", "Mobile Phone",
    "Fund", "Gross", "Fee", "Payment Method", "Date", "Deposit Date",
    "Batch Name", "Notes", "Batch Number", "Check Number"
]

NAME_EXCLUDE_WORDS = [
    "bank", "credit union", "savings", "checking", "national", "federal",
    "pay to", "order of", "dollars", "memo", "for", "void", "date",
    "routing", "account", "address", "street", "ave", "blvd", "rd",
    "apt", "suite", "po box", "p.o.", "city", "state", "zip",
]


# -- Contacts Lookup ---------------------------------------------------------

def load_contacts(contacts_path):
    """Load contacts from a CHMeetings contacts export xlsx.
    Expected headers: Name, First Name, Last Name, ...
    Returns list of dicts with 'first_name' and 'last_name'."""
    wb = load_workbook(contacts_path, read_only=True, data_only=True)
    ws = wb.active

    # Find header row and column indices
    headers = {}
    for col, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=False))):
        val = str(cell.value or "").strip().lower()
        headers[val] = col

    fn_col = headers.get("first name")
    ln_col = headers.get("last name")

    if fn_col is None or ln_col is None:
        print(f"   WARNING: Contacts file missing 'First Name' or 'Last Name' columns.")
        print(f"   Found columns: {[str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]}")
        wb.close()
        return []

    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        first = str(row[fn_col] or "").strip()
        last = str(row[ln_col] or "").strip()
        if first or last:
            contacts.append({"first_name": first, "last_name": last})

    wb.close()
    print(f"   Loaded {len(contacts)} contacts.")
    return contacts


def match_contact(first_name, last_name, contacts, threshold=0.6):
    """Find the best matching contact for a given name.
    Uses fuzzy matching on the combined full name.
    Returns (first_name, last_name, score) of best match, or None."""
    if not contacts or (not first_name and not last_name):
        return None

    ocr_full = f"{first_name} {last_name}".lower().strip()
    best_match = None
    best_score = 0

    for contact in contacts:
        contact_full = f"{contact['first_name']} {contact['last_name']}".lower().strip()

        # Full name similarity
        score = difflib.SequenceMatcher(None, ocr_full, contact_full).ratio()

        # Also try last-name-only matching (weighted higher since last names are more distinctive)
        if last_name and contact['last_name']:
            ln_score = difflib.SequenceMatcher(
                None, last_name.lower(), contact['last_name'].lower()
            ).ratio()
            # If last name is a strong match, boost the overall score
            score = max(score, ln_score * 0.9)

        if score > best_score:
            best_score = score
            best_match = contact

    if best_match and best_score >= threshold:
        return (best_match['first_name'], best_match['last_name'], best_score)

    return None


# -- Autocomplete Input ------------------------------------------------------

def _get_name_suggestions(partial, contacts, field="first_name"):
    """Return contacts whose field starts with `partial` (case-insensitive)."""
    if not partial:
        return []
    p = partial.lower()
    seen = set()
    results = []
    for c in contacts:
        val = c.get(field, "")
        if val.lower().startswith(p) and val.lower() not in seen:
            seen.add(val.lower())
            results.append(c)
    return results[:10]  # cap at 10


def input_with_autocomplete(prompt, default, contacts, field="first_name"):
    """Read user input with Tab-autocomplete against the contacts list.

    Works on Windows (msvcrt) and Unix (readline).
    Falls back to plain input() if neither is available.
    """
    if not contacts:
        new_val = input(f"  {prompt} [{default}]: ").strip()
        return new_val if new_val else default

    # --- Try Windows msvcrt approach ---
    try:
        import msvcrt
        return _input_autocomplete_msvcrt(prompt, default, contacts, field)
    except ImportError:
        pass

    # --- Try readline approach (Linux / Mac / WSL) ---
    try:
        import readline

        completions = []

        def completer(text, state):
            nonlocal completions
            if state == 0:
                matches = _get_name_suggestions(text, contacts, field)
                completions = [c[field] for c in matches]
            return completions[state] if state < len(completions) else None

        old_completer = readline.get_completer()
        old_delims = readline.get_completer_delims()
        readline.set_completer(completer)
        readline.set_completer_delims("")  # complete on entire input
        readline.parse_and_bind("tab: complete")
        try:
            new_val = input(f"  {prompt} [{default}] (Tab=autocomplete): ").strip()
        finally:
            readline.set_completer(old_completer)
            readline.set_completer_delims(old_delims)

        return new_val if new_val else default

    except ImportError:
        pass

    # --- Plain fallback ---
    new_val = input(f"  {prompt} [{default}]: ").strip()
    return new_val if new_val else default


def _input_autocomplete_msvcrt(prompt, default, contacts, field):
    """Character-by-character input with inline autocomplete on Windows."""
    import msvcrt

    sys.stdout.write(f"  {prompt} [{default}] (Tab=autocomplete): ")
    sys.stdout.flush()

    buf = []
    suggestion = ""

    while True:
        ch = msvcrt.getwch()

        if ch == '\r':  # Enter
            sys.stdout.write('\n')
            result = "".join(buf).strip()
            return result if result else default

        elif ch == '\t':  # Tab - accept suggestion or cycle
            matches = _get_name_suggestions("".join(buf), contacts, field)
            if matches:
                # Accept first match
                chosen = matches[0][field]
                # Clear current line and rewrite
                sys.stdout.write('\r' + ' ' * 120 + '\r')
                buf = list(chosen)
                sys.stdout.write(f"  {prompt} [{default}] (Tab=autocomplete): {chosen}")
                sys.stdout.flush()

        elif ch == '\x08' or ch == '\x7f':  # Backspace
            if buf:
                buf.pop()
                # Redraw line
                sys.stdout.write('\r' + ' ' * 120 + '\r')
                current = "".join(buf)
                hint = ""
                matches = _get_name_suggestions(current, contacts, field) if current else []
                if matches:
                    hint = f"  -> {', '.join(c[field] for c in matches[:3])}"
                sys.stdout.write(f"  {prompt} [{default}] (Tab=autocomplete): {current}{hint}")
                sys.stdout.flush()

        elif ch == '\x1b':  # Escape - clear
            buf = []
            sys.stdout.write('\r' + ' ' * 120 + '\r')
            sys.stdout.write(f"  {prompt} [{default}] (Tab=autocomplete): ")
            sys.stdout.flush()

        elif ch >= ' ':  # Normal printable character
            buf.append(ch)
            current = "".join(buf)
            matches = _get_name_suggestions(current, contacts, field)
            hint = ""
            if matches:
                hint = f"  -> {', '.join(c[field] for c in matches[:3])}"
            # Redraw line
            sys.stdout.write('\r' + ' ' * 120 + '\r')
            sys.stdout.write(f"  {prompt} [{default}] (Tab=autocomplete): {current}{hint}")
            sys.stdout.flush()


# -- Image Display -----------------------------------------------------------

class CheckImageViewer:
    """Shows check images in a Tkinter window that stays open while
    the user enters data in the terminal."""

    def __init__(self):
        if not HAS_TK:
            self.enabled = False
            return
        self.enabled = True
        self.root = None
        self.label = None
        self.photo = None
        self._ready = threading.Event()
        self._thread = threading.Thread(target=self._run_tk, daemon=True)
        self._thread.start()
        self._ready.wait(timeout=5)

    def _run_tk(self):
        self.root = tk.Tk()
        self.root.title("Check Viewer")
        self.root.configure(bg="gray20")
        self.label = tk.Label(self.root, bg="gray20")
        self.label.pack(padx=10, pady=10)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self._ready.set()
        self.root.mainloop()

    def _on_close(self):
        """User closed the window — disable further updates."""
        self.enabled = False
        self.root.destroy()

    def show_check(self, image_bytes, title=""):
        """Display a check image. Call from main thread."""
        if not self.enabled or not self.root:
            return
        try:
            img = Image.open(io.BytesIO(image_bytes))
            # Scale to fit nicely on screen (max 900px wide)
            max_w = 900
            if img.width > max_w:
                ratio = max_w / img.width
                img = img.resize((max_w, int(img.height * ratio)), Image.LANCZOS)
            self.root.after(0, self._update_image, img, title)
        except Exception:
            pass

    def _update_image(self, img, title):
        try:
            self.photo = ImageTk.PhotoImage(img)
            self.label.configure(image=self.photo)
            if title:
                self.root.title(title)
        except Exception:
            pass

    def close(self):
        if self.enabled and self.root:
            try:
                self.root.after(0, self.root.destroy)
            except Exception:
                pass


# -- Image Preprocessing ----------------------------------------------------

def preprocess_check_image(image_bytes):
    img = Image.open(io.BytesIO(image_bytes))
    w, h = img.size
    if w < 1500:
        scale = 2.0
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    img = img.convert("L")
    img = img.filter(ImageFilter.SHARPEN)
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    enhancer = ImageEnhance.Brightness(img)
    img = enhancer.enhance(1.2)
    return np.array(img)


# -- OCR Engine --------------------------------------------------------------

class CheckOCR:
    def __init__(self, languages=None):
        langs = languages or ['en']
        print("Loading OCR engine (first run downloads ~100MB model)...")
        self.reader = easyocr.Reader(langs, gpu=False)

    def extract_text_blocks(self, image_bytes):
        img_array = preprocess_check_image(image_bytes)
        results = self.reader.readtext(img_array, detail=1, paragraph=False)
        return results


# -- Region-Based Check Parser -----------------------------------------------

def classify_block_region(bbox, img_width, img_height):
    center_x = (bbox[0][0] + bbox[2][0]) / 2
    center_y = (bbox[0][1] + bbox[2][1]) / 2
    if center_y < img_height * 0.33:
        vert = "top"
    elif center_y < img_height * 0.66:
        vert = "middle"
    else:
        vert = "bottom"
    horiz = "left" if center_x < img_width * 0.55 else "right"
    return f"{vert}_{horiz}"


def is_likely_name(text):
    text_lower = text.lower().strip()
    if len(text_lower) < 3 or len(text_lower) > 50:
        return False
    for word in NAME_EXCLUDE_WORDS:
        if word in text_lower:
            return False
    if re.match(r'^[\d\s.,/$#*-]+$', text):
        return False
    alpha_chars = sum(1 for c in text if c.isalpha())
    if alpha_chars < len(text.strip()) * 0.6:
        return False
    return True


def extract_first_last(text):
    """Extract just first and last name, dropping middle initials/names.
    'RON S BOOMER' -> ('Ron', 'Boomer')
    'JUDY K SCHNEIDER' -> ('Judy', 'Schneider')
    'MICHAEL SCHNEIDER' -> ('Michael', 'Schneider')
    'SUSAN J CHASTAIN' -> ('Susan', 'Chastain')
    """
    parts = text.split()
    if len(parts) == 0:
        return "", ""
    if len(parts) == 1:
        return "", parts[0].title()
    # First word is always the first name, last word is always the last name
    # Everything in between is middle name/initial — drop it
    return parts[0].title(), parts[-1].title()


def parse_check_text(text_blocks, img_width, img_height, verbose=False):
    """Parse OCR text blocks using spatial region analysis.
    Returns best-guess name, check number, and amount."""
    all_text = " ".join([t[1] for t in text_blocks])

    regions = {}
    for block in text_blocks:
        bbox, text, conf = block
        region = classify_block_region(bbox, img_width, img_height)
        if region not in regions:
            regions[region] = []
        regions[region].append({
            "text": text.strip(), "confidence": conf, "bbox": bbox,
            "center_y": (bbox[0][1] + bbox[2][1]) / 2,
            "center_x": (bbox[0][0] + bbox[2][0]) / 2,
        })

    for region in regions:
        regions[region].sort(key=lambda b: (b["center_y"], b["center_x"]))

    if verbose:
        print("\n   --- OCR TEXT BY REGION ---")
        for region in sorted(regions.keys()):
            print(f"   [{region}]")
            for b in regions[region]:
                print(f"     ({b['confidence']:.2f}) {b['text']}")
        print("   --- END OCR TEXT ---\n")

    result = {"first_name": "", "last_name": "", "check_number": "", "amount": "", "raw_text": all_text}

    # -- NAME: top-left, first + last only (drop middle initials/names) --
    top_left = regions.get("top_left", [])
    for block in top_left:
        if is_likely_name(block["text"]):
            first, last = extract_first_last(block["text"])
            if first and last:
                result["first_name"] = first
                result["last_name"] = last
            elif last:
                result["last_name"] = last
            break

    if not result["first_name"] and len(top_left) >= 2:
        first_two = [b["text"] for b in top_left[:2]]
        if all(is_likely_name(t) for t in first_two):
            combined = " ".join(first_two)
            first, last = extract_first_last(combined)
            if first and last:
                result["first_name"] = first
                result["last_name"] = last

    # -- CHECK NUMBER: top-right --
    top_right = regions.get("top_right", [])
    for block in top_right:
        text = block["text"].strip()
        m = re.match(r'^[#]?(\d{3,6})$', text)
        if m:
            result["check_number"] = m.group(1)
            break
        m = re.search(r'(?:No\.?|Check\s*#?|Ck\s*#?|#)\s*(\d{3,6})', text, re.IGNORECASE)
        if m:
            result["check_number"] = m.group(1)
            break

    if not result["check_number"]:
        for block in text_blocks:
            m = re.search(r'(?:No\.?|Check\s*#?|Ck\s*#?)\s*(\d{3,6})', block[1].strip(), re.IGNORECASE)
            if m:
                result["check_number"] = m.group(1)
                break

    # -- AMOUNT: prefer right side --
    # OCR often splits amounts across blocks (e.g. "$", "250", "OO")
    # Strategy: 1) try reconstructing from blocks near "$" sign
    #           2) try concatenated region text
    #           3) try individual blocks
    #           4) fallback to all text
    amounts = []

    def fix_ocr_zeros(text):
        """Fix OO -> 00 (letter O misread as digit 0)."""
        text = re.sub(r'\bOO\b', '00', text)
        text = re.sub(r'\b[Oo][Oo]\b', '00', text)
        return text

    # Strategy 1: Find "$" block in middle-right, gather nearby digit blocks
    mr_blocks = regions.get("middle_right", [])
    dollar_block = None
    for b in mr_blocks:
        if '$' in b["text"]:
            dollar_block = b
            break

    if dollar_block:
        # Gather all purely numeric parts from middle-right blocks
        # then reconstruct as $[dollars].[cents]
        numeric_parts = []
        for b in mr_blocks:
            t = fix_ocr_zeros(b["text"].strip())
            # Extract digit-only tokens
            digits_only = re.sub(r'[^0-9]', '', t)
            if digits_only:
                numeric_parts.append(digits_only)

        if numeric_parts:
            # Separate into dollars part (longer numbers) and cents part (exactly 2 digits)
            dollars_parts = []
            cents_part = "00"
            for p in numeric_parts:
                if len(p) == 2 and int(p) < 100:
                    cents_part = p  # likely the cents
                elif len(p) >= 1:
                    dollars_parts.append(p)

            if dollars_parts:
                # Use the largest as the dollar amount
                dollar_val = max(dollars_parts, key=lambda x: int(x))
                reconstructed = f"${dollar_val}.{cents_part}"
                for amt in extract_amounts(reconstructed):
                    amounts.append({"value": amt, "priority": 0})

    # Strategy 2: concatenated middle-right text
    mr_text = fix_ocr_zeros(" ".join([b["text"] for b in mr_blocks]))
    for amt in extract_amounts(mr_text):
        amounts.append({"value": amt, "priority": 1})

    # Strategy 3: individual blocks from priority regions
    priority_blocks = mr_blocks + regions.get("top_right", [])
    for block in priority_blocks:
        text_fixed = fix_ocr_zeros(block["text"])
        for amt in extract_amounts(text_fixed):
            amounts.append({"value": amt, "priority": 1})

    # Strategy 4: anywhere in the document
    all_fixed = fix_ocr_zeros(all_text)
    for amt in extract_amounts(all_fixed):
        amounts.append({"value": amt, "priority": 2})

    if amounts:
        # Prefer highest priority (lowest number), then largest value
        best_priority = min(a["priority"] for a in amounts)
        top_tier = [a for a in amounts if a["priority"] == best_priority]
        best = max(top_tier, key=lambda a: a["value"])
        result["amount"] = f"{best['value']:.2f}"

    return result


def extract_amounts(text):
    amounts = []
    patterns = [
        r'\$\s*([\d,]+\.\d{2})',                   # $1,234.56
        r'\$\s*([\d]+)[,.](\d{2})\b',             # $70,00 or $70.00
        r'\$\s*([\d,]+)\s*\.\s*(\d{2})',           # $ 70 . 00
        r'\$\s*([\d,]+)\s+(\d{2})\b',             # $ 250 00 (space instead of decimal, from split OCR blocks)
        r'\*+\s*([\d,]+[.,]\d{2})\s*\*+',         # **1,234.56**
        r'(?<=\s)([\d]+)[.,](\d{2})(?=\s|$)',     # standalone 150.00 or 150,00
    ]
    for pattern in patterns:
        for m in re.finditer(pattern, text):
            try:
                groups = m.groups()
                if len(groups) == 1:
                    val = float(groups[0].replace(",", ""))
                elif len(groups) == 2:
                    whole = groups[0].replace(",", "").replace(" ", "")
                    val = float(f"{whole}.{groups[1]}")
                else:
                    continue
                if 0.01 <= val <= 999999.99:
                    amounts.append(val)
            except ValueError:
                continue
    return amounts


# -- PDF Processing ----------------------------------------------------------

def extract_check_images_from_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    images = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        image_list = page.get_images(full=True)
        if image_list:
            for img_index, img_info in enumerate(image_list):
                xref = img_info[0]
                base_image = doc.extract_image(xref)
                if base_image:
                    images.append({"bytes": base_image["image"], "page": page_num + 1, "index": img_index + 1})
        else:
            mat = fitz.Matrix(3, 3)
            pix = page.get_pixmap(matrix=mat)
            images.append({"bytes": pix.tobytes("png"), "page": page_num + 1, "index": 1})
    doc.close()
    return images


# -- Interactive Review with Image Display -----------------------------------

def review_entry(entry, index, total, image_bytes=None, viewer=None, contacts=None):
    """Show the check image and let user review/correct the extracted data."""

    # Display check image in viewer window
    if viewer and image_bytes:
        viewer.show_check(image_bytes, title=f"Check {index}/{total}")

    # Check for contact match
    contact_match = None
    if contacts and (entry.get('first_name') or entry.get('last_name')):
        contact_match = match_contact(entry['first_name'], entry['last_name'], contacts)

    print(f"\n{'='*60}")
    print(f"  Check {index}/{total}  (Page {entry.get('_page', '?')})")
    print(f"  >> Check image is displayed in the viewer window <<")
    print(f"{'='*60}")
    print(f"  First Name:   {entry['first_name'] or '(empty)'}")
    print(f"  Last Name:    {entry['last_name'] or '(empty)'}")
    if contact_match:
        score_pct = int(contact_match[2] * 100)
        print(f"  -> Match:     {contact_match[0]} {contact_match[1]}  ({score_pct}% match)")
    print(f"  Check #:      {entry['check_number'] or '(empty)'}")
    print(f"  Amount:       ${entry['amount'] or '(empty)'}")
    print(f"  Memo:         {entry['notes'] or '(none)'}")
    print(f"  Fund:         {entry['fund'] or '(none)'}")
    print(f"{'='*60}")

    options = "  [A]ccept"
    if contact_match:
        options += " / [M]atch (use contact name)"
    options += " / [E]dit / [S]kip / [P]revious?"

    choice = input(options).strip().lower()

    if choice == 'p':
        return "PREVIOUS"
    elif choice == 's':
        return None
    elif choice == 'm' and contact_match:
        entry['first_name'] = contact_match[0]
        entry['last_name'] = contact_match[1]
        print(f"  -> Using: {entry['first_name']} {entry['last_name']}")
        return entry
    elif choice == 'e':
        print("  (Press Enter to keep current value; Tab to autocomplete names)")
        # Name fields get autocomplete against contacts list
        entry['first_name'] = input_with_autocomplete(
            "First Name", entry['first_name'], contacts or [], "first_name"
        )
        entry['last_name'] = input_with_autocomplete(
            "Last Name", entry['last_name'], contacts or [], "last_name"
        )
        # Other fields use plain input
        for field, label in [
            ("check_number", "Check #"),
            ("amount", "Amount"),
            ("notes", "Memo"),
            ("fund", "Fund"),
        ]:
            new_val = input(f"  {label} [{entry[field]}]: ").strip()
            if new_val:
                entry[field] = new_val
        return entry
    else:
        return entry


# -- Excel Output ------------------------------------------------------------

def write_to_excel(entries, output_path, batch_name="", batch_number="", deposit_date=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contributions"

    for col, header in enumerate(CHMEETINGS_COLUMNS, 1):
        ws.cell(row=1, column=col, value=header)

    for row_num, entry in enumerate(entries, 2):
        ws.cell(row=row_num, column=1, value=entry.get("first_name", ""))
        ws.cell(row=row_num, column=2, value=entry.get("last_name", ""))
        ws.cell(row=row_num, column=3, value=entry.get("envelope_number", ""))
        ws.cell(row=row_num, column=4, value=entry.get("email", ""))
        ws.cell(row=row_num, column=5, value=entry.get("mobile_phone", ""))
        ws.cell(row=row_num, column=6, value=entry.get("fund", ""))
        ws.cell(row=row_num, column=7, value=entry.get("amount", ""))
        ws.cell(row=row_num, column=8, value=entry.get("fee", ""))
        ws.cell(row=row_num, column=9, value=entry.get("payment_method", "Check"))
        ws.cell(row=row_num, column=10, value=entry.get("date", ""))
        ws.cell(row=row_num, column=11, value=deposit_date)
        ws.cell(row=row_num, column=12, value=batch_name)
        ws.cell(row=row_num, column=13, value=entry.get("notes", ""))
        ws.cell(row=row_num, column=14, value=batch_number)
        ws.cell(row=row_num, column=15, value=entry.get("check_number", ""))

    for col in range(1, len(CHMEETINGS_COLUMNS) + 1):
        max_len = len(CHMEETINGS_COLUMNS[col - 1])
        for row in range(2, len(entries) + 2):
            cell_val = str(ws.cell(row=row, column=col).value or "")
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 3

    wb.save(output_path)
    return output_path


# -- Main --------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Extract check data from a PDF and create a CHMeetings import file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python check_to_chmeetings.py "checks.pdf" --review --contacts "contacts.xlsx"
  python check_to_chmeetings.py "checks.pdf" --review --contacts "contacts.xlsx" --fund "Tithes"
  python check_to_chmeetings.py "checks.pdf" --review --fund "Tithes" --date "03/08/2026"
  python check_to_chmeetings.py "checks.pdf" --review --batch "March 2026" --deposit-date "03/08/2026"

In --review mode, each check image is displayed in a popup window so you
can read the actual check while confirming or correcting the extracted data.
        """
    )
    parser.add_argument("pdf", help="Path to the PDF containing check images")
    parser.add_argument("-o", "--output", help="Output Excel file path")
    parser.add_argument("--review", action="store_true", help="Interactive review mode with image display")
    parser.add_argument("--verbose", action="store_true", help="Show all OCR text by region (for debugging)")
    parser.add_argument("--contacts", default="", help="Path to CHMeetings contacts export xlsx for name matching")
    parser.add_argument("--fund", default="", help="Default fund name (e.g. 'Tithes')")
    parser.add_argument("--date", default="", help="Contribution date for all entries (MM/DD/YYYY)")
    parser.add_argument("--batch", default="", help="Batch name for this import")
    parser.add_argument("--batch-number", default="", help="Batch number for this import")
    parser.add_argument("--deposit-date", default="", help="Deposit date (MM/DD/YYYY)")
    parser.add_argument("--payment-method", default="Check", help="Payment method (default: Check)")

    args = parser.parse_args()

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        sys.exit(f"Error: PDF not found: {pdf_path}")

    if not args.output:
        today = datetime.now().strftime("%Y-%m-%d")
        args.output = f"contributions_{today}.xlsx"

    # Step 1: Extract images from PDF
    print(f"\nReading PDF: {pdf_path}")
    check_images = extract_check_images_from_pdf(str(pdf_path))
    print(f"   Found {len(check_images)} image(s) in the PDF.")

    if not check_images:
        sys.exit("No images found in the PDF.")

    # Step 2: Load contacts if provided
    contacts = []
    if args.contacts:
        contacts_path = Path(args.contacts)
        if contacts_path.exists():
            print(f"\nLoading contacts: {contacts_path}")
            contacts = load_contacts(str(contacts_path))
        else:
            print(f"\nWARNING: Contacts file not found: {contacts_path}")

    # Step 3: Start image viewer if in review mode
    viewer = None
    if args.review and HAS_TK:
        print("   Opening check image viewer...")
        viewer = CheckImageViewer()
    elif args.review and not HAS_TK:
        print("   Note: tkinter not available, image display disabled.")

    # Step 3: OCR each image
    ocr = CheckOCR()
    entries = []          # final accepted entries, indexed by check position
    entry_map = {}        # position -> entry (so we can go back and edit)
    parsed_cache = {}     # position -> parsed OCR result (avoid re-OCR)
    skipped = set()       # positions that were skipped

    i = 0
    total = len(check_images)
    while i < total:
        img_data = check_images[i]

        # Use cached OCR if we already processed this check (going back)
        if i in parsed_cache:
            parsed = parsed_cache[i]
        else:
            print(f"\nProcessing check {i+1}/{total} (page {img_data['page']})...")
            text_blocks = ocr.extract_text_blocks(img_data["bytes"])
            if not text_blocks:
                print(f"   WARNING: No text detected in image {i+1}, skipping.")
                i += 1
                continue

            img = Image.open(io.BytesIO(img_data["bytes"]))
            img_w, img_h = img.size
            if img_w < 1500:
                img_w *= 2
                img_h *= 2

            parsed = parse_check_text(text_blocks, img_w, img_h, verbose=args.verbose)
            parsed_cache[i] = parsed

        # Use previously edited entry if going back, otherwise build from OCR
        if i in entry_map:
            entry = entry_map[i]
        else:
            entry = {
                "first_name": parsed["first_name"],
                "last_name": parsed["last_name"],
                "envelope_number": "",
                "email": "",
                "mobile_phone": "",
                "fund": args.fund,
                "amount": parsed["amount"],
                "fee": "",
                "payment_method": args.payment_method,
                "date": args.date,
                "notes": "",
                "check_number": parsed["check_number"],
                "_page": img_data["page"],
                "_raw_text": parsed["raw_text"],
            }

            # Auto-apply strong contact match when not revisiting an edited entry
            if contacts:
                cm = match_contact(entry['first_name'], entry['last_name'], contacts, threshold=0.75)
                if cm:
                    entry['first_name'] = cm[0]
                    entry['last_name'] = cm[1]

        if args.review:
            result = review_entry(entry, i + 1, total,
                                  image_bytes=img_data["bytes"], viewer=viewer,
                                  contacts=contacts)
            if result == "PREVIOUS":
                if i > 0:
                    i -= 1
                    print("   Going back to previous check...")
                else:
                    print("   Already at the first check.")
                continue
            elif result is None:
                skipped.add(i)
                entry_map.pop(i, None)
                print("   Skipped.")
                i += 1
                continue
            else:
                entry = result

        # Store the entry
        entry_map[i] = entry
        skipped.discard(i)

        name = f"{entry['first_name']} {entry['last_name']}".strip() or "(no name)"
        amt = f"${entry['amount']}" if entry['amount'] else "(no amount)"
        ck = f"#{entry['check_number']}" if entry['check_number'] else "(no check #)"
        memo = f" [{entry['notes']}]" if entry['notes'] else ""

        # Running total from all accepted entries so far
        all_accepted = [entry_map[k] for k in sorted(entry_map.keys()) if k not in skipped]
        running_total = sum(float(e['amount']) for e in all_accepted if e.get('amount'))
        print(f"   -> {name} -- {amt} -- {ck}{memo}")
        print(f"   Running total: ${running_total:,.2f}  ({len(all_accepted)} checks)")

        i += 1

    # Build final entries list in order
    entries = [entry_map[k] for k in sorted(entry_map.keys()) if k not in skipped]

    # Close viewer
    if viewer:
        viewer.close()

    if not entries:
        sys.exit("\nNo entries extracted. Nothing to export.")

    # Summary
    total_amount = sum(float(e['amount']) for e in entries if e.get('amount'))
    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"{'='*60}")
    print(f"  Total checks:    {len(entries)}")
    print(f"  Total amount:    ${total_amount:,.2f}")
    print(f"{'='*60}")

    # Step 4: Write Excel
    print(f"\nWriting {len(entries)} entries to: {args.output}")
    write_to_excel(
        entries, args.output,
        batch_name=args.batch,
        batch_number=args.batch_number,
        deposit_date=args.deposit_date,
    )

    print(f"\nDone! Import into CHMeetings under Contributions -> Import.")

    # Save debug JSON
    debug_path = Path(args.output).stem + "_debug.json"
    debug_data = []
    for e in entries:
        d = {k: v for k, v in e.items() if not k.startswith("_")}
        d["_raw_text"] = e.get("_raw_text", "")
        d["_page"] = e.get("_page", "")
        debug_data.append(d)

    with open(debug_path, "w") as f:
        json.dump(debug_data, f, indent=2)
    print(f"   Debug data saved to: {debug_path}")


if __name__ == "__main__":
    main()
